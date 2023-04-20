import openpyxl

# Your excel file path
path = "XXXXXXXX"

wb = openpyxl.load_workbook(path)
# can change the active sheet if multiple sheets
wb.active = wb['Sheet1']
sheet = wb.active

max_col = sheet.max_column
max_row = sheet.max_row

desired_row = int(input("Enter the desired row  \n"))
sel_row = input("Wanna Before (or) After entered row (B/A) \n").upper()

if sel_row == "A":
    for i in range(desired_row, max_row + 1):
        print("\n")
        for j in range(1, max_col + 1):
            cell = sheet.cell(row=i, column=j)
            print(cell.value, end=" | ")

elif sel_row == "B":
    upper_rev = int((desired_row + max_row) / 2)
    #  for printing from the reverse order use the line below
    # for i in range(desired_row, 1, -1):
    for i in range(1, desired_row + 1):
        print("\n")
        for j in range(1, max_col + 1):
            cell = sheet.cell(row=i, column=j)
            print(cell.value, end=" | ")

else:
    print("\n")
    print("OOPS! Invalid data Re-run again")
